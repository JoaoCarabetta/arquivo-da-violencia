#!/usr/bin/env python3
"""
Load official violence data from Ministry of Justice bancovde-YYYY.xlsx into the database.

This script downloads and ingests municipal victim counts by month from the
SINESP VDE (Validador de Dados Estatísticos) open data portal.

Data source: https://www.gov.br/mj/pt-br/assuntos/sua-seguranca/seguranca-publica/estatistica/download/dnsp-base-de-dados/
URL pattern: bancovde-YYYY.xlsx

Usage:
    python scripts/load_official_violence_data.py [--year YYYY] [--since YYYY-MM]

Options:
    --year YYYY      Year to download (default: 2025)
    --since YYYY-MM  Only load data from this month onwards (default: 2025-09)
"""

import asyncio
import sys
from pathlib import Path
from io import BytesIO

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from sqlmodel import SQLModel
from sqlmodel.ext.asyncio.session import AsyncSession
from loguru import logger

from app.database import get_engine, init_db
from app.services.official_violence_data import ingest_official_violence_data

# bancovde URL pattern (gov.br portal)
BANCOVDE_URL_TEMPLATE = "https://www.gov.br/mj/pt-br/assuntos/sua-seguranca/seguranca-publica/estatistica/download/dnsp-base-de-dados/bancovde-{year}.xlsx/@@download/file"

# Expected bancovde-2025.xlsx headers (14 columns)
EXPECTED_HEADERS = [
    "uf", "municipio", "evento", "data_referencia", "agente", "arma",
    "faixa_etaria", "feminino", "masculino", "nao_informado",
    "total_vitima", "total", "total_peso", "abrangencia"
]

async def download_and_parse_bancovde(year: int = 2025, since_year_month: str = "2025-09") -> list:
    """
    Download and parse bancovde-YYYY.xlsx file.

    Expected structure:
    - Sheet name: "{year}" (e.g. "2025")
    - 14 columns (see EXPECTED_HEADERS)
    - ~832k rows per year
    - Excel serial dates in data_referencia column

    Municipality resolution:
    - No 7-digit IBGE code in file
    - Resolve (uf, municipio) → code_muni via ibge_population table
    - Drop rows that don't match

    Args:
        year: Year to download (e.g. 2025)
        since_year_month: Only return data >= this month (YYYY-MM format)

    Returns:
        List of parsed bancovde data rows (dicts with 14 columns)
    """
    url = BANCOVDE_URL_TEMPLATE.format(year=year)
    logger.info(f"Downloading bancovde-{year}.xlsx from {url}...")

    try:
        import httpx
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from datetime import datetime, timedelta

        # Download XLSX
        async with httpx.AsyncClient(timeout=120.0) as client:
            response = await client.get(url)
            response.raise_for_status()

        logger.info(f"Downloaded {len(response.content)} bytes")

        # Load workbook
        wb = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        sheet_name = str(year)

        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in workbook. "
                f"Available sheets: {wb.sheetnames}"
            )

        ws = wb[sheet_name]
        logger.info(f"Loaded sheet '{sheet_name}' with {ws.max_row} rows")

        # Read header row (row 1)
        headers = []
        for col_idx in range(1, 15):  # 14 columns (A-N)
            cell = ws.cell(row=1, column=col_idx)
            header_value = cell.value
            if header_value is None:
                # Empty header - use positional fallback
                headers.append(f"col_{col_idx}")
            else:
                headers.append(str(header_value).strip())

        logger.info(f"Headers: {headers}")

        # Verify headers match expected structure
        if headers != EXPECTED_HEADERS:
            logger.warning(
                f"Header mismatch! Expected: {EXPECTED_HEADERS}, "
                f"Got: {headers}"
            )

        # Parse data rows (skip header row)
        records = []
        row_count = 0
        skipped_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_count += 1

            # Parse row into dict (14 columns)
            record = {}
            for col_idx, value in enumerate(row):
                if col_idx >= len(headers):
                    break
                header = headers[col_idx]
                record[header] = value

            # Filter by date
            data_ref = record.get("data_referencia")
            if data_ref is not None:
                try:
                    # Convert Excel serial date to YYYY-MM
                    excel_epoch = datetime(1899, 12, 30)
                    date = excel_epoch + timedelta(days=int(data_ref))
                    year_month = date.strftime("%Y-%m")

                    # Skip rows before cutoff
                    if year_month < since_year_month:
                        skipped_count += 1
                        continue
                except (ValueError, TypeError):
                    # Invalid date - skip
                    skipped_count += 1
                    continue
            else:
                # Missing date - skip
                skipped_count += 1
                continue

            records.append(record)

            # Progress log every 100k rows
            if row_count % 100000 == 0:
                logger.info(f"Processed {row_count} rows, kept {len(records)}")

        wb.close()

        logger.info(f"Loaded {len(records)} rows from bancovde-{year}.xlsx (skipped {skipped_count} rows)")
        logger.info(f"Filtered to {len(records)} rows >= {since_year_month}")

        return records

    except ImportError as e:
        logger.error(f"Missing required package: {e}. Install with: pip install httpx openpyxl")
        raise
    except Exception as e:
        logger.error(f"Failed to download/parse bancovde-{year}.xlsx: {e}")
        raise

async def main(year: int = 2025, since: str = "2025-09"):
    """Load official violence data into database."""
    logger.info("=" * 80)
    logger.info(f"Loading official violence data (bancovde-{year}.xlsx, since={since})")
    logger.info("=" * 80)

    # Initialize database tables
    await init_db()
    logger.info("Database tables verified")

    # Get engine
    engine = get_engine()

    # Download and parse bancovde data
    try:
        vde_data = await download_and_parse_bancovde(year=year, since_year_month=since)
    except Exception as e:
        logger.error(f"✗ Failed to download bancovde data: {e}")
        logger.info("Manual alternative: Download the file from www.gov.br and adapt this script")
        return

    # Ingest data
    async with AsyncSession(engine) as session:
        try:
            await ingest_official_violence_data(
                session=session,
                vde_data=vde_data
            )
            logger.success("✓ Official violence data loaded successfully")
            logger.info("=" * 80)
            logger.info("Next steps:")
            logger.info("  1. Verify: SELECT COUNT(*) FROM official_violence_count;")
            logger.info("  2. Test coverage calculation (issue 176)")
            logger.info("=" * 80)
        except Exception as e:
            logger.error(f"✗ Failed to ingest official violence data: {e}")
            raise

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Load official violence data from Ministry of Justice bancovde-YYYY.xlsx"
    )
    parser.add_argument(
        "--year",
        type=int,
        default=2025,
        help="Year to download (default: 2025)"
    )
    parser.add_argument(
        "--since",
        type=str,
        default="2025-09",
        help="Only load data from this month onwards (YYYY-MM format, default: 2025-09)"
    )

    args = parser.parse_args()

    asyncio.run(main(year=args.year, since=args.since))
